# -*- coding: utf-8 -*-
"""
导入导出模块 - 处理数据的导入和导出
支持CSV和Excel格式(xls/xlsx)
"""
import wx
import wx.adv as wxadv
import csv
from wx.lib.agw.aquabutton import AquaButton
from database import Database
from i18n import get_text, add_language_listener, remove_language_listener
from i18n_support import LanguageSupportMixin
from theme import get_theme, get_button_colors, get_card_colors, add_theme_listener

# 尝试导入Excel支持库
try:
    import openpyxl
    from openpyxl import Workbook
    EXCEL_SUPPORT = True
except ImportError:
    try:
        import xlwt
        import xlrd
        EXCEL_SUPPORT = True
    except ImportError:
        EXCEL_SUPPORT = False


def read_excel_file(file_path):
    """读取Excel文件"""
    try:
        # 尝试使用openpyxl (xlsx格式)
        wb = openpyxl.load_workbook(file_path)
        ws = wb.active
        data = []
        for row in ws.iter_rows(values_only=True):
            data.append([str(cell) if cell is not None else '' for cell in row])
        return data
    except:
        return None


def write_excel_file(file_path, headers, rows):
    """写入Excel文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "收支记录"
    
    # 写入表头
    ws.append(headers)
    
    # 写入数据
    for row in rows:
        ws.append(row)
    
    wb.save(file_path)


class ImportDialog(wx.Dialog, LanguageSupportMixin):
    """导入对话框"""
    
    def __init__(self, parent=None, user_id=None, db=None):
        self.user_id = user_id
        self.db = db or Database()
        self.file_path = None

        super().__init__(parent, title=get_text('导入数据'), size=(500, 350))
        
        # 设置窗口图标
        try:
            self.SetIcon(wx.Icon("family_accounting.png", wx.BITMAP_TYPE_PNG))
        except:
            pass
        
        self.init_ui()
        self.setup_language_listener()
        self.Centre()
        
        # 添加主题切换监听
        add_theme_listener(self.on_theme_changed)
        # 应用主题
        self.apply_theme()
    
    def on_theme_changed(self, theme_name):
        """主题切换回调"""
        if not self or not self.IsShown():
            return
        wx.CallAfter(self.apply_theme)
    
    def apply_theme(self):
        """应用当前主题"""
        from theme import apply_theme_to_window
        from config_manager import config
        theme = get_theme()
        self.SetBackgroundColour(theme['bg_color'])
        # 递归应用主题到所有控件（包括按钮）
        apply_theme_to_window(self, config.get_theme())
        self.Refresh()
    
    def refresh_labels(self):
        """刷新界面文字"""
        self.SetTitle(get_text('导入数据'))
        if EXCEL_SUPPORT:
            hint_text = f"{get_text('从CSV/Excel文件导入收支记录')}\n{get_text('支持格式')}: CSV, XLS, XLSX"
        else:
            hint_text = f"{get_text('从CSV文件导入收支记录')}\n{get_text('支持格式')}: CSV"
        self.hint.SetLabel(hint_text)
        self.browse_btn.SetLabel(get_text('选择文件导入'))
        self.import_btn.SetLabel(get_text('导入'))
        self.cancel_btn.SetLabel(get_text('取消'))
        
        # 更新预览列表列标题
        self.preview_list.SetColumnTitle(0, get_text('日期'))
        self.preview_list.SetColumnTitle(1, get_text('类型'))
        self.preview_list.SetColumnTitle(2, get_text('分类'))
        self.preview_list.SetColumnTitle(3, get_text('金额'))
        self.preview_list.SetColumnTitle(4, get_text('说明'))
        self.preview_list.SetColumnTitle(5, get_text('货币'))

    def init_ui(self):
        """初始化界面"""
        panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        # 提示信息
        if EXCEL_SUPPORT:
            hint_text = f"{get_text('从CSV/Excel文件导入收支记录')}\n{get_text('支持格式')}: CSV, XLS, XLSX"
        else:
            hint_text = f"{get_text('从CSV文件导入收支记录')}\n{get_text('支持格式')}: CSV"

        self.hint = wx.StaticText(panel, label=hint_text)
        main_sizer.Add(self.hint, flag=wx.ALL, border=15)

        # 文件选择
        file_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.file_text = wx.TextCtrl(panel, size=(300, -1), style=wx.TE_READONLY)
        self.browse_btn = AquaButton(panel, label=get_text('选择文件导入'), size=(120, 28))
        self.browse_btn.SetForegroundColour(wx.Colour(0, 0, 0))
        self.browse_btn.Bind(wx.EVT_BUTTON, self.on_browse)
        file_sizer.Add(self.file_text, proportion=1)
        file_sizer.Add(self.browse_btn, flag=wx.LEFT, border=10)
        main_sizer.Add(file_sizer, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=20)
        main_sizer.Add(wx.StaticText(panel, label=""), flag=wx.EXPAND, border=10)

        # 预览
        preview_label = wx.StaticText(panel, label=get_text('预览') + ":")
        self.preview_list = wx.ListCtrl(panel, size=(460, 150), style=wx.LC_REPORT)
        self.preview_list.InsertColumn(0, get_text('日期'), width=90)
        self.preview_list.InsertColumn(1, get_text('类型'), width=50)
        self.preview_list.InsertColumn(2, get_text('分类'), width=70)
        self.preview_list.InsertColumn(3, get_text('金额'), width=80)
        self.preview_list.InsertColumn(4, get_text('说明'), width=100)
        self.preview_list.InsertColumn(5, get_text('货币'), width=50)

        main_sizer.Add(preview_label, flag=wx.LEFT, border=20)
        main_sizer.Add(self.preview_list, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=20)
        main_sizer.Add(wx.StaticText(panel, label=""), flag=wx.EXPAND, border=10)

        # 按钮
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.import_btn = AquaButton(panel, label=get_text('导入'), size=(90, 30))
        self.cancel_btn = AquaButton(panel, label=get_text('取消'), size=(90, 30))
        
        # 设置按钮文字为黑色
        btn_fg = wx.Colour(0, 0, 0)
        for btn in [self.import_btn, self.cancel_btn]:
            btn.SetForegroundColour(btn_fg)
        
        self.import_btn.Bind(wx.EVT_BUTTON, self.on_import)
        self.cancel_btn.Bind(wx.EVT_BUTTON, self.on_cancel)
        btn_sizer.Add(self.import_btn)
        btn_sizer.Add(self.cancel_btn, flag=wx.LEFT, border=20)
        main_sizer.Add(btn_sizer, flag=wx.ALIGN_CENTER | wx.BOTTOM, border=15)

        panel.SetSizer(main_sizer)

    def on_browse(self, event):
        """浏览文件"""
        if EXCEL_SUPPORT:
            wildcard = f"{get_text('所有支持格式')} (*.csv;*.xls;*.xlsx)|*.csv;*.xls;*.xlsx|CSV (*.csv)|*.csv|Excel (*.xlsx)|*.xlsx|Excel 97-2003 (*.xls)|*.xls"
        else:
            wildcard = "CSV (*.csv)|*.csv"

        dlg = wx.FileDialog(self, get_text('选择文件'), wildcard=wildcard,
                           style=wx.FD_OPEN | wx.FD_FILE_MUST_EXIST)
        if dlg.ShowModal() == wx.ID_OK:
            self.file_path = dlg.GetPath()
            self.file_text.SetValue(self.file_path)
            self.preview_file()
        dlg.Destroy()

    def preview_file(self):
        """预览文件"""
        self.preview_list.DeleteAllItems()
        if not self.file_path:
            return

        try:
            ext = self.file_path.lower().split('.')[-1]

            if ext in ('xls', 'xlsx'):
                if not EXCEL_SUPPORT:
                    wx.MessageBox(get_text('Excel支持库未安装'), get_text('错误'), wx.OK | wx.ICON_ERROR)
                    return
                data = read_excel_file(self.file_path)
                if data is None:
                    wx.MessageBox(get_text('读取文件失败'), get_text('错误'), wx.OK | wx.ICON_ERROR)
                    return
                rows = data
            else:
                with open(self.file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)

            for i, row in enumerate(rows):
                if i >= 5:
                    break
                if len(row) >= 4:
                    index = self.preview_list.InsertItem(0, row[0] if len(row) > 0 else '')
                    self.preview_list.SetItem(index, 1, row[1] if len(row) > 1 else '')
                    self.preview_list.SetItem(index, 2, row[2] if len(row) > 2 else '')
                    self.preview_list.SetItem(index, 3, row[3] if len(row) > 3 else '')
                    self.preview_list.SetItem(index, 4, row[4] if len(row) > 4 else '')
                    self.preview_list.SetItem(index, 5, row[5] if len(row) > 5 else 'CNY')
        except Exception as e:
            wx.MessageBox(f"{get_text('读取文件失败')}: {str(e)}", get_text('错误'), wx.OK | wx.ICON_ERROR)

    def on_import(self, event):
        """导入数据"""
        if not self.file_path:
            wx.MessageBox(get_text('请先选择文件'), get_text('提示'), wx.OK | wx.ICON_WARNING)
            return

        try:
            imported = 0
            skipped = 0
            ext = self.file_path.lower().split('.')[-1]

            if ext in ('xls', 'xlsx'):
                if not EXCEL_SUPPORT:
                    wx.MessageBox(get_text('Excel支持库未安装'), get_text('错误'), wx.OK | wx.ICON_ERROR)
                    return
                data = read_excel_file(self.file_path)
                if data is None:
                    wx.MessageBox(get_text('读取文件失败'), get_text('错误'), wx.OK | wx.ICON_ERROR)
                    return
                rows = data[1:] if len(data) > 0 else []  # 跳过表头
            else:
                with open(self.file_path, 'r', encoding='utf-8-sig') as f:
                    reader = csv.reader(f)
                    rows = list(reader)

            for row in rows:
                if len(row) < 4:
                    skipped += 1
                    continue

                date = row[0].strip()
                type_ = row[1].strip()
                category = row[2].strip()

                try:
                    amount = float(row[3].strip())
                except ValueError:
                    skipped += 1
                    continue

                description = row[4].strip() if len(row) > 4 else ''

                if type_ not in ('收入', '支出'):
                    skipped += 1
                    continue

                currency = row[5].strip() if len(row) > 5 else 'CNY'

                success, _ = self.db.add_account(
                    self.user_id, category, type_, amount, category, description, date, currency
                )
                if success:
                    imported += 1
                else:
                    skipped += 1

            msg = f"{get_text('导入成功')}!\n{get_text('成功')}: {imported}\n{get_text('跳过')}: {skipped}"
            wx.MessageBox(msg, get_text('导入完成'), wx.OK | wx.ICON_INFORMATION)

            if imported > 0:
                self.EndModal(wx.ID_OK)

        except Exception as e:
            wx.MessageBox(f"{get_text('导入失败')}: {str(e)}", get_text('错误'), wx.OK | wx.ICON_ERROR)

    def on_cancel(self, event):
        """取消"""
        self.EndModal(wx.ID_CANCEL)


class ExportDialog(wx.Dialog, LanguageSupportMixin):
    """导出对话框"""
    
    def __init__(self, parent=None, user_id=None, db=None):
        self.user_id = user_id
        self.db = db or Database()
        super().__init__(parent, title=get_text('导出数据'), size=(400, 330))
        
        # 设置窗口图标
        try:
            self.SetIcon(wx.Icon("family_accounting.png", wx.BITMAP_TYPE_PNG))
        except:
            pass
        
        self.init_ui()
        self.setup_language_listener()
        self.Centre()
        
        # 添加主题切换监听
        add_theme_listener(self.on_theme_changed)
        # 应用主题
        self.apply_theme()
    
    def on_theme_changed(self, theme_name):
        """主题切换回调"""
        if not self or not self.IsShown():
            return
        wx.CallAfter(self.apply_theme)
    
    def apply_theme(self):
        """应用当前主题"""
        from theme import apply_theme_to_window
        from config_manager import config
        theme = get_theme()
        self.SetBackgroundColour(theme['bg_color'])
        # 递归应用主题到所有控件（包括按钮）
        apply_theme_to_window(self, config.get_theme())
        self.Refresh()
    
    def refresh_labels(self):
        """刷新界面文字"""
        self.SetTitle(get_text('导出数据'))
        self.format_box.SetLabel(get_text('导出格式'))
        self.date_range_box.SetLabel(get_text('日期筛选'))
        self.start_label.SetLabel(f"{get_text('开始日期')}:")
        self.end_label.SetLabel(f"{get_text('结束日期')}:")
        self.export_btn.SetLabel(get_text('导出'))
        self.cancel_btn.SetLabel(get_text('取消'))

    def init_ui(self):
        """初始化界面"""
        panel = wx.Panel(self)
        main_sizer = wx.BoxSizer(wx.VERTICAL)

        hint = wx.StaticText(panel, label=get_text('将收支记录导出为CSV/Excel文件'))
        main_sizer.Add(hint, flag=wx.ALL, border=15)

        # 导出格式选择
        self.format_box = wx.StaticBox(panel, label=get_text('导出格式'))
        format_sizer = wx.StaticBoxSizer(self.format_box, wx.VERTICAL)

        self.format_choice = wx.Choice(panel, choices=['CSV (*.csv)', 'Excel (*.xlsx)', 'Excel 97-2003 (*.xls)'])
        self.format_choice.SetSelection(0)
        format_sizer.Add(self.format_choice, flag=wx.ALL, border=5)
        main_sizer.Add(format_sizer, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=20)
        main_sizer.Add(wx.StaticText(panel, label=""), flag=wx.EXPAND, border=10)

        # 日期范围
        self.date_range_box = wx.StaticBox(panel, label=get_text('日期筛选'))
        date_range_sizer = wx.StaticBoxSizer(self.date_range_box, wx.VERTICAL)

        start_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.start_label = wx.StaticText(panel, label=f"{get_text('开始日期')}:")
        self.start_picker = wxadv.DatePickerCtrl(panel)
        start_sizer.Add(self.start_label, flag=wx.ALIGN_CENTER_VERTICAL)
        start_sizer.Add(self.start_picker, flag=wx.LEFT, border=10)
        date_range_sizer.Add(start_sizer, flag=wx.ALL, border=5)

        end_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.end_label = wx.StaticText(panel, label=f"{get_text('结束日期')}:")
        self.end_picker = wxadv.DatePickerCtrl(panel)
        self.end_picker.SetValue(wx.DateTime.Now())
        end_sizer.Add(self.end_label, flag=wx.ALIGN_CENTER_VERTICAL)
        end_sizer.Add(self.end_picker, flag=wx.LEFT, border=10)
        date_range_sizer.Add(end_sizer, flag=wx.ALL, border=5)

        main_sizer.Add(date_range_sizer, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=20)
        main_sizer.Add(wx.StaticText(panel, label=""), flag=wx.EXPAND, border=10)

        # 按钮
        btn_sizer = wx.BoxSizer(wx.HORIZONTAL)
        self.export_btn = AquaButton(panel, label=get_text('导出'), size=(120, 30))
        self.cancel_btn = AquaButton(panel, label=get_text('取消'), size=(120, 30))
        
        # 设置按钮文字为黑色
        btn_fg = wx.Colour(0, 0, 0)
        for btn in [self.export_btn, self.cancel_btn]:
            btn.SetForegroundColour(btn_fg)
        
        self.export_btn.Bind(wx.EVT_BUTTON, self.on_export)
        self.cancel_btn.Bind(wx.EVT_BUTTON, self.on_cancel)
        btn_sizer.Add(self.export_btn)
        btn_sizer.Add(self.cancel_btn, flag=wx.LEFT, border=20)
        main_sizer.Add(btn_sizer, flag=wx.ALIGN_CENTER | wx.BOTTOM, border=15)

        panel.SetSizer(main_sizer)
    
    def on_export(self, event):
        """导出数据"""
        export_type = self.format_choice.GetSelection()

        if export_type == 0:
            wildcard = "CSV (*.csv)|*.csv"
            default_ext = ".csv"
        elif export_type == 1:
            wildcard = "Excel (*.xlsx)|*.xlsx"
            default_ext = ".xlsx"
        else:
            wildcard = "Excel 97-2003 (*.xls)|*.xls"
            default_ext = ".xls"

        dlg = wx.FileDialog(self, get_text('选择文件导入'), wildcard=wildcard,
                           style=wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
        if dlg.ShowModal() == wx.ID_OK:
            file_path = dlg.GetPath()

            # 确保文件扩展名正确
            if not file_path.lower().endswith(default_ext):
                file_path += default_ext

            start_date = self.start_picker.GetValue().Format('%Y-%m-%d')
            end_date = self.end_picker.GetValue().Format('%Y-%m-%d')

            accounts = self.db.get_accounts(self.user_id, include_deleted=False)
            filtered = [a for a in accounts if start_date <= a['date'] <= end_date]

            try:
                headers = [get_text('日期'), get_text('类型'), get_text('分类'), get_text('金额'), get_text('说明'), get_text('货币')]
                rows = []
                for acc in filtered:
                    rows.append([
                        acc['date'],
                        acc['type'],
                        acc.get('category', ''),
                        acc['amount'],
                        acc.get('description', ''),
                        acc.get('currency', 'CNY')
                    ])

                if export_type == 0:
                    # CSV导出
                    with open(file_path, 'w', encoding='utf-8-sig', newline='') as f:
                        writer = csv.writer(f)
                        writer.writerow(headers)
                        writer.writerows(rows)
                elif export_type == 1:
                    # Excel导出 (xlsx)
                    if not EXCEL_SUPPORT:
                        wx.MessageBox(get_text('Excel支持库未安装'), get_text('错误'), wx.OK | wx.ICON_ERROR)
                        dlg.Destroy()
                        return
                    write_excel_file(file_path, headers, rows)
                else:
                    wx.MessageBox("xls format requires xlwt library", get_text('提示'), wx.OK | wx.ICON_INFORMATION)

                msg = f"{get_text('导出成功')}!\n{get_text('共')} {len(filtered)} {get_text('条记录')}\n{get_text('保存至')}: {file_path}"
                wx.MessageBox(msg, get_text('导出完成'), wx.OK | wx.ICON_INFORMATION)
                self.EndModal(wx.ID_OK)

            except Exception as e:
                wx.MessageBox(f"{get_text('导出失败')}: {str(e)}", get_text('错误'), wx.OK | wx.ICON_ERROR)

        dlg.Destroy()

    def on_cancel(self, event):
        """取消"""
        self.EndModal(wx.ID_CANCEL)


class ImportExportPanel(wx.Panel):
    """导入导出面板"""
    
    def __init__(self, parent, user_id, db):
        super().__init__(parent)
        self.user_id = user_id
        self.db = db
        self.init_ui()
        
        # 添加语言切换监听
        add_language_listener(self.on_language_changed)
    
    def on_language_changed(self):
        """语言切换回调"""
        try:
            if self and not self.IsBeingDeleted():
                wx.CallAfter(self.refresh_labels)
        except:
            pass
    
    def apply_theme(self):
        """应用当前主题"""
        theme = get_theme()
        
        # 设置面板背景
        self.SetBackgroundColour(theme['bg_color'])
        
        # 刷新
        self.Refresh()
    
    def init_ui(self):
        """初始化界面"""
        main_sizer = wx.BoxSizer(wx.VERTICAL)
        
        # 标题
        self.title_label = wx.StaticText(self, label=get_text('数据导入导出'))
        title_font = wx.Font(14, wx.FONTFAMILY_DEFAULT, wx.FONTSTYLE_NORMAL, wx.FONTWEIGHT_BOLD)
        self.title_label.SetFont(title_font)
        main_sizer.Add(self.title_label, flag=wx.ALL, border=15)
        
        # 导入卡片
        self.import_box = wx.StaticBox(self, label=get_text('导入数据'))
        import_sizer = wx.StaticBoxSizer(self.import_box, wx.VERTICAL)
        
        if EXCEL_SUPPORT:
            self.import_desc = wx.StaticText(self, label=get_text('从CSV/Excel文件导入收支记录') + "\n" + "CSV, XLS, XLSX")
        else:
            self.import_desc = wx.StaticText(self, label=get_text('从CSV文件导入收支记录') + "\n" + "CSV")
        import_sizer.Add(self.import_desc, flag=wx.ALL, border=10)
        
        self.import_btn = AquaButton(self, label=get_text('选择文件导入'), size=(150, 35))
        self.import_btn.SetForegroundColour(wx.Colour(0, 0, 0))
        self.import_btn.Bind(wx.EVT_BUTTON, self.on_import)
        import_sizer.Add(self.import_btn, flag=wx.ALL | wx.ALIGN_CENTER, border=10)
        
        main_sizer.Add(import_sizer, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=20)
        main_sizer.Add(wx.StaticText(self, label=""), flag=wx.EXPAND, border=15)
        
        # 导出卡片
        self.export_box = wx.StaticBox(self, label=get_text('导出数据'))
        export_sizer = wx.StaticBoxSizer(self.export_box, wx.VERTICAL)
        
        if EXCEL_SUPPORT:
            self.export_desc = wx.StaticText(self, label=get_text('将收支记录导出为CSV/Excel文件'))
        else:
            self.export_desc = wx.StaticText(self, label=get_text('将收支记录导出为CSV文件'))
        export_sizer.Add(self.export_desc, flag=wx.ALL, border=10)
        
        self.export_btn = AquaButton(self, label=get_text('导出'), size=(150, 35))
        self.export_btn.SetForegroundColour(wx.Colour(0, 0, 0))
        self.export_btn.Bind(wx.EVT_BUTTON, self.on_export)
        export_sizer.Add(self.export_btn, flag=wx.ALL | wx.ALIGN_CENTER, border=10)
        
        main_sizer.Add(export_sizer, flag=wx.EXPAND | wx.LEFT | wx.RIGHT, border=20)
        
        self.SetSizer(main_sizer)
    
    def refresh_labels(self):
        """刷新标签文字"""
        self.title_label.SetLabel(get_text('数据导入导出'))
        self.import_box.SetLabel(get_text('导入数据'))
        self.export_box.SetLabel(get_text('导出数据'))
        self.import_btn.SetLabel(get_text('选择文件导入'))
        self.export_btn.SetLabel(get_text('导出'))
        if EXCEL_SUPPORT:
            self.import_desc.SetLabel(get_text('从CSV/Excel文件导入收支记录') + "\n" + "CSV, XLS, XLSX")
            self.export_desc.SetLabel(get_text('将收支记录导出为CSV/Excel文件'))
        else:
            self.import_desc.SetLabel(get_text('从CSV文件导入收支记录') + "\n" + "CSV")
            self.export_desc.SetLabel(get_text('将收支记录导出为CSV文件'))
        self.Layout()
    
    def on_import(self, event):
        """导入"""
        dlg = ImportDialog(self, self.user_id, self.db)
        if dlg.ShowModal() == wx.ID_OK:
            self.GetParent().refresh_accounts()
        dlg.Destroy()
    
    def on_export(self, event):
        """导出"""
        dlg = ExportDialog(self, self.user_id, self.db)
        dlg.ShowModal()
        dlg.Destroy()


if __name__ == "__main__":
    app = wx.App()
    frame = wx.Frame(None)
    db = Database()
    panel = ImportExportPanel(frame, user_id=1, db=db)
    frame.Show()
    app.MainLoop()
